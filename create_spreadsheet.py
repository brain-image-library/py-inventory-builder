import os
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# ─────────────────────────────────────────────────────────────
# Setup paths and date strings
# ─────────────────────────────────────────────────────────────
tsv_path = "summary_metadata.tsv"
year_month = datetime.now().strftime("%Y%m")          # e.g., 202508
excel_path = f"{year_month}.xlsx"                     # e.g., 202508.xlsx
sheet_name = datetime.now().strftime("%Y%m%d")        # e.g., 20250806

# ─────────────────────────────────────────────────────────────
# Load TSV into a DataFrame
# ─────────────────────────────────────────────────────────────
df = pd.read_csv(tsv_path, sep="\t")

# ─────────────────────────────────────────────────────────────
# Sort rows by score (ascending)
# ─────────────────────────────────────────────────────────────
if 'score' in df.columns:
    df = df.sort_values(by='score', ascending=True)

# ─────────────────────────────────────────────────────────────
# Check if Excel file is corrupted and delete if so
# ─────────────────────────────────────────────────────────────
if os.path.exists(excel_path):
    try:
        _ = load_workbook(excel_path)
    except Exception as e:
        print(f"⚠️ Warning: '{excel_path}' is invalid: {e}")
        print("🗑️ Deleting corrupted Excel file and starting fresh.")
        os.remove(excel_path)

# ─────────────────────────────────────────────────────────────
# Check if sheet already exists in workbook
# ─────────────────────────────────────────────────────────────
sheet_exists = False
if os.path.exists(excel_path):
    try:
        wb = load_workbook(excel_path)
        sheet_exists = sheet_name in wb.sheetnames
    except Exception as e:
        print(f"⚠️ Warning: Unable to open workbook '{excel_path}' to check sheet: {e}")

# ─────────────────────────────────────────────────────────────
# Write to Excel only if sheet doesn't exist
# ─────────────────────────────────────────────────────────────
if not sheet_exists:
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a' if os.path.exists(excel_path) else 'w') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ─────────────────────────────────────────────────────────
    # Format and style the new worksheet
    # ─────────────────────────────────────────────────────────
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    ws.freeze_panes = "A2"

    # Style header
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.value = str(cell.value).upper()
        cell.fill = header_fill
        cell.font = header_font

    # Format score column and highlight rows with score < 1
    if 'score' in df.columns:
        score_col_idx = list(df.columns).index('score') + 1
        percent_format = '0.0%'
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            score_cell = row[score_col_idx - 1]
            try:
                if score_cell.value is not None:
                    score_cell.number_format = percent_format
                    if float(score_cell.value) < 1:
                        for cell in row:
                            cell.fill = yellow_fill
            except ValueError:
                continue

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Save workbook
    wb.save(excel_path)
else:
    print(f"ℹ️ Sheet '{sheet_name}' already exists in '{excel_path}', skipping creation and formatting.")

# ─────────────────────────────────────────────────────────────
# Upload the Excel file to PSC via rclone
# ─────────────────────────────────────────────────────────────
try:
    subprocess.run(
        ["rclone", "copy", excel_path, "PSC:Brain_Image_Library/reports/daily/"],
        check=True
    )
    print(f"✅ Uploaded {excel_path} to PSC:Brain_Image_Library/reports/daily/")
except subprocess.CalledProcessError as e:
    print(f"❌ rclone upload failed: {e}")
